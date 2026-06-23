"""
Helmet & Plate Detection System
Pipeline: two_wheeler -> ROI -> helmet -> (no-helmet only) -> plate -> OCR
- Video:          GPU YOLO + voting buffer + frame skip
- Image/Webcam:   GPU YOLO + immediate OCR log (no voting)
Saves to violations.txt and violations.xlsx
"""

import re
import time
import threading
import tkinter as tk
from tkinter import filedialog
from PIL import Image, ImageTk
import cv2
import numpy as np
from ultralytics import YOLO
import easyocr
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

# ── Models ─────────────────────────────────────────────────────────────────────
print("[INFO] Loading models...")
two_wheeler_model = YOLO("models/two_wheeler.pt").to("cuda")
helmet_model      = YOLO("models/helmet.pt").to("cuda")
plate_model       = YOLO("models/plate.pt").to("cuda")
ocr               = easyocr.Reader(['en'], gpu=True, verbose=False)
print("[INFO] Models loaded.")

# ── Inference sizes ────────────────────────────────────────────────────────────
TW_IMGSZ    = 640
HELM_IMGSZ  = 512
PLATE_IMGSZ = 640

# ── Detection thresholds ───────────────────────────────────────────────────────
TW_CONF       = 0.25
TW_IOU        = 0.40
HELMET_CONF   = 0.20
PLATE_CONF    = 0.20
EXPAND_W      = 0.20
EXPAND_H_UP   = 1.20
EXPAND_H_DOWN = 0.25

# ── Box sanity filters ─────────────────────────────────────────────────────────
TW_MIN_W_RATIO = 0.08
TW_MAX_W_RATIO = 0.92
TW_MIN_H_RATIO = 0.12

# ── Class names ────────────────────────────────────────────────────────────────
NO_HELMET_NAMES = {"no-helmet", "no helmet", "nohelmet"}

# ── Colors BGR ────────────────────────────────────────────────────────────────
COLOR_TW     = (30,  200,  30)
COLOR_NOHELM = (0,     0, 230)
COLOR_PLATE  = (230,  60,   0)
COLOR_ROI    = (0,  220, 255)

# ── Output files ───────────────────────────────────────────────────────────────
VIOLATIONS_TXT  = "number plate log/violations.txt"
VIOLATIONS_XLSX = "number plate log/violations.xlsx"

# ── Video-specific settings ────────────────────────────────────────────────────
FRAME_SKIP      = 2    # process every Nth frame for video
OCR_VOTE_WINDOW = 4    # how many matching reads before committing plate

# ── Dedup ─────────────────────────────────────────────────────────────────────
PLATE_COOLDOWN  = 8.0  # seconds before same plate can log again

# ── Global state ───────────────────────────────────────────────────────────────
running          = False
source_type      = "camera"
source_path      = None
cap              = None
plate_log        = []
detection_thread = None
current_frame    = None
frame_lock       = threading.Lock()
_seen_plates     = {}
_ocr_buffer      = {}   # used only during video
_last_photo      = None


# ══════════════════════════════════════════════════════════════════════════════
# Helpers shared by both pipelines
# ══════════════════════════════════════════════════════════════════════════════

def correct_plate(text):
    t = text.replace(" ", "").upper()
    if len(t) < 4:
        return text
    TO_LETTER = {'0':'O','1':'I','2':'Z','5':'S','6':'G','8':'B'}
    TO_DIGIT  = {'O':'0','I':'1','Z':'2','S':'5','G':'6','B':'8','Q':'0','D':'0'}

    m = re.match(r'^([A-Z0-9]{2})([0-9]{2})([A-Z]{1,2})([0-9]{3,4})$', t)
    if m:
        p1 = ''.join(TO_LETTER.get(c, c) for c in m.group(1))
        p2 = m.group(2)
        p3 = m.group(3)
        p4 = m.group(4)
        return f"{p1} {p2} {p3} {p4}"

    # Try with corrections first then rematch
    corrected = []
    for i, c in enumerate(t):
        if i < 2:
            corrected.append(TO_LETTER.get(c, c))
        elif i < 4:
            corrected.append(TO_DIGIT.get(c, c))
        elif i < 6:
            corrected.append(TO_LETTER.get(c, c))
        else:
            corrected.append(TO_DIGIT.get(c, c))
    t2 = ''.join(corrected)

    m2 = re.match(r'^([A-Z]{2})([0-9]{2})([A-Z]{1,2})([0-9]{3,4})$', t2)
    if m2:
        return f"{m2.group(1)} {m2.group(2)} {m2.group(3)} {m2.group(4)}"

    # Fallback: just space as XX XX rest
    if len(t2) >= 4:
        return t2[:2] + ' ' + t2[2:4] + ' ' + t2[4:]
    return t2

def clean_plate(raw):
    t = raw.upper()
    t = re.sub(r"[^A-Z0-9]", " ", t)
    t = re.sub(r" +", " ", t).strip()
    t = correct_plate(t)
    return t


def is_valid_plate(text):
    t = text.replace(" ", "").upper()
    if len(t) < 6 or len(t) > 13:
        return False
    if not any(c.isdigit() for c in t):
        return False
    if not any(c.isalpha() for c in t):
        return False
    if sum(c.isdigit() for c in t) < 2:
        return False
    if sum(c.isalpha() for c in t) < 2:
        return False
    if max(t.count(c) for c in set(t)) / len(t) > 0.55:
        return False
    return True


def edit_distance(a, b):
    a, b = a.replace(" ", ""), b.replace(" ", "")
    dp = list(range(len(b) + 1))
    for i, ca in enumerate(a):
        ndp = [i + 1]
        for j, cb in enumerate(b):
            ndp.append(min(dp[j] + (ca != cb), dp[j+1] + 1, ndp[j] + 1))
        dp = ndp
    return dp[len(b)]


def expand_box(x1, y1, x2, y2, fh, fw):
    bw, bh  = x2 - x1, y2 - y1
    up_px   = max(int(bh * EXPAND_H_UP),   int(fh * 0.25))
    down_px = max(int(bh * EXPAND_H_DOWN), 30)
    side_px = max(int(bw * EXPAND_W),      20)
    return (max(0, x1 - side_px),
            max(0, y1 - up_px),
            min(fw - 1, x2 + side_px),
            min(fh - 1, y2 + down_px))


def safe_crop(img, x1, y1, x2, y2):
    if x2 <= x1 or y2 <= y1:
        return None
    c = img[y1:y2, x1:x2]
    return c if c.size > 0 else None


PLATE_NOISE = {"POLICE", "FOLICE", "POLCE", "POUCE", "P0LICE", "GOVT", "IND"}

def run_ocr(crop):
    try:
        gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        gray = clahe.apply(gray)
        sharp_kernel = np.array([[0, -1, 0], [-1, 5, -1], [0, -1, 0]], np.float32)
        gray = cv2.filter2D(gray, -1, sharp_kernel)
        processed = cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR)
        results = ocr.readtext(
            processed, detail=1, paragraph=False,
            allowlist='ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789 ',
            text_threshold=0.6, low_text=0.4
        )
        if not results:
            return ""
        parts = [t.strip() for (_, t, c) in results
                 if c >= 0.25 and t.strip()
                 and t.strip().upper() not in PLATE_NOISE]
        return " ".join(parts)
    except Exception as e:
        print(f"[WARN] OCR: {e}")
        return ""


def draw_label(img, text, x1, y1, color):
    font, fs, th = cv2.FONT_HERSHEY_SIMPLEX, 0.55, 1
    (tw, txh), bl = cv2.getTextSize(text, font, fs, th)
    top = max(y1 - txh - bl - 4, 0)
    cv2.rectangle(img, (x1, top), (x1 + tw + 6, top + txh + bl + 6), color, -1)
    cv2.putText(img, text, (x1 + 3, top + txh + 3),
                font, fs, (255, 255, 255), th, cv2.LINE_AA)


def upscale_frame(frame):
    fh0, fw0 = frame.shape[:2]
    if fw0 < 1000:
        sc = 1280 / fw0
        frame = cv2.resize(frame, (int(fw0*sc), int(fh0*sc)),
                           interpolation=cv2.INTER_LINEAR)
    return frame


def detect_two_wheelers(orig, fh, fw):
    """Run two-wheeler model and return list of valid (x1,y1,x2,y2) boxes."""
    try:
        res = two_wheeler_model(orig, conf=TW_CONF, iou=TW_IOU,
                                imgsz=TW_IMGSZ, device="cuda",
                                verbose=False)[0]
    except Exception as e:
        print(f"[WARN] TW: {e}")
        return []
    if res.boxes is None:
        return []
    boxes = []
    for box in res.boxes:
        try:
            x1, y1, x2, y2 = map(int, box.xyxy[0].tolist())
        except Exception:
            continue
        bw, bh = x2 - x1, y2 - y1
        if bw < fw * TW_MIN_W_RATIO: continue
        if bh < fh * TW_MIN_H_RATIO: continue
        if bw > fw * TW_MAX_W_RATIO: continue
        boxes.append((x1, y1, x2, y2))
    return boxes


def detect_no_helmet(roi):
    """Return (found: bool, best_box_in_roi, conf)."""
    try:
        res = helmet_model(roi, conf=HELMET_CONF,
                           imgsz=HELM_IMGSZ, device="cuda",
                           verbose=False)[0]
    except Exception as e:
        print(f"[WARN] Helmet: {e}")
        return False, None, 0.0
    if res.boxes is None:
        return False, None, 0.0
    best_box, best_conf = None, -1.0
    found = False
    for hb in res.boxes:
        try:
            cid   = int(hb.cls[0])
            cname = helmet_model.names[cid].strip().lower()
            hconf = float(hb.conf[0])
        except Exception:
            continue
        if cname in NO_HELMET_NAMES:
            found = True
            if hconf > best_conf:
                best_conf = hconf
                try:
                    best_box = tuple(map(int, hb.xyxy[0].tolist()))
                except Exception:
                    pass
    return found, best_box, best_conf


def detect_plate(roi):
    """Return best plate box (x1,y1,x2,y2) in ROI coords or None."""
    try:
        res = plate_model(roi, conf=PLATE_CONF,
                          imgsz=PLATE_IMGSZ, device="cuda",
                          verbose=False)[0]
    except Exception as e:
        print(f"[WARN] Plate: {e}")
        return None
    if res.boxes is None or len(res.boxes) == 0:
        return None
    best_pb, best_pc = None, -1.0
    for pb in res.boxes:
        try:
            pc = float(pb.conf[0])
        except Exception:
            continue
        if pc > best_pc:
            best_pc, best_pb = pc, pb
    if best_pb is None:
        return None
    try:
        px1, py1, px2, py2 = map(int, best_pb.xyxy[0].tolist())
    except Exception:
        return None
    if (px2 - px1) < 10 or (py2 - py1) < 5:
        return None
    return (px1, py1, px2, py2)


def get_plate_crop(orig, rx1, ry1, rx2, ry2, px1, py1, px2, py2, fw, fh):
    PAD  = 8
    fpx1 = max(0,      rx1 + px1 - PAD)
    fpy1 = max(0,      ry1 + py1 - PAD)
    fpx2 = min(fw - 1, rx1 + px2 + PAD)
    fpy2 = min(fh - 1, ry1 + py2 + PAD)
    crop = safe_crop(orig, fpx1, fpy1, fpx2, fpy2)
    if crop is not None:
        ph, pw = crop.shape[:2]
        target_w = max(400, pw)
        if pw < target_w:
            sc = target_w / pw
            crop = cv2.resize(crop, (int(pw*sc), int(ph*sc)),
                              interpolation=cv2.INTER_CUBIC)
    return crop, fpx1, fpy1, fpx2, fpy2


# ══════════════════════════════════════════════════════════════════════════════
# IMAGE & WEBCAM pipeline  — immediate OCR, no voting
# ══════════════════════════════════════════════════════════════════════════════

def process_frame_live(frame):
    """
    Used for IMAGE and WEBCAM.
    Detects, runs OCR immediately, returns (annotated_frame, [plate_strings]).
    """
    frame  = upscale_frame(frame)
    orig   = frame.copy()
    canvas = frame.copy()
    fh, fw = frame.shape[:2]
    found_plates = []

    tw_boxes = detect_two_wheelers(orig, fh, fw)

    for (x1, y1, x2, y2) in tw_boxes:
        cv2.rectangle(canvas, (x1, y1), (x2, y2), COLOR_TW, 2)
        draw_label(canvas, "Two-Wheeler", x1, y1, COLOR_TW)

        rx1, ry1, rx2, ry2 = expand_box(x1, y1, x2, y2, fh, fw)
        cv2.rectangle(canvas, (rx1, ry1), (rx2, ry2), COLOR_ROI, 1)
        roi = safe_crop(orig, rx1, ry1, rx2, ry2)
        if roi is None:
            continue

        found, nh_box, nh_conf = detect_no_helmet(roi)
        if not found:
            continue

        if nh_box:
            hx1, hy1, hx2, hy2 = nh_box
            cv2.rectangle(canvas,
                          (rx1+hx1, ry1+hy1), (rx1+hx2, ry1+hy2),
                          COLOR_NOHELM, 2)
            draw_label(canvas, f"No Helmet {nh_conf:.0%}",
                       rx1+hx1, ry1+hy1, COLOR_NOHELM)

        plate_box = detect_plate(roi)
        if plate_box is None:
            continue

        px1, py1, px2, py2 = plate_box
        crop, fpx1, fpy1, fpx2, fpy2 = get_plate_crop(
            orig, rx1, ry1, rx2, ry2, px1, py1, px2, py2, fw, fh)

        # Always draw the plate bounding box
        cv2.rectangle(canvas, (fpx1, fpy1), (fpx2, fpy2), COLOR_PLATE, 2)

        if crop is None:
            draw_label(canvas, "Plate", fpx1, fpy1, COLOR_PLATE)
            continue

        raw = run_ocr(crop)
        if raw:
            cleaned = clean_plate(raw)
            if is_valid_plate(cleaned):
                draw_label(canvas, cleaned, fpx1, fpy1, COLOR_PLATE)
                found_plates.append(cleaned)
            else:
                draw_label(canvas, "Plate", fpx1, fpy1, COLOR_PLATE)
        else:
            draw_label(canvas, "Plate", fpx1, fpy1, COLOR_PLATE)

    return canvas, found_plates


# ══════════════════════════════════════════════════════════════════════════════
# VIDEO pipeline  — frame skip + voting buffer
# ══════════════════════════════════════════════════════════════════════════════

def process_frame_video(frame):
    """
    Used for VIDEO only.
    Uses voting buffer — returns found_plates only when vote threshold reached.
    """
    global _ocr_buffer

    frame  = upscale_frame(frame)
    orig   = frame.copy()
    canvas = frame.copy()
    fh, fw = frame.shape[:2]
    found_plates = []

    tw_boxes = detect_two_wheelers(orig, fh, fw)

    for (x1, y1, x2, y2) in tw_boxes:
        cv2.rectangle(canvas, (x1, y1), (x2, y2), COLOR_TW, 2)
        draw_label(canvas, "Two-Wheeler", x1, y1, COLOR_TW)

        rx1, ry1, rx2, ry2 = expand_box(x1, y1, x2, y2, fh, fw)
        cv2.rectangle(canvas, (rx1, ry1), (rx2, ry2), COLOR_ROI, 1)
        roi = safe_crop(orig, rx1, ry1, rx2, ry2)
        if roi is None:
            continue

        found, nh_box, nh_conf = detect_no_helmet(roi)
        if not found:
            continue

        if nh_box:
            hx1, hy1, hx2, hy2 = nh_box
            cv2.rectangle(canvas,
                          (rx1+hx1, ry1+hy1), (rx1+hx2, ry1+hy2),
                          COLOR_NOHELM, 2)
            draw_label(canvas, f"No Helmet {nh_conf:.0%}",
                       rx1+hx1, ry1+hy1, COLOR_NOHELM)

        plate_box = detect_plate(roi)
        if plate_box is None:
            continue

        px1, py1, px2, py2 = plate_box
        crop, fpx1, fpy1, fpx2, fpy2 = get_plate_crop(
            orig, rx1, ry1, rx2, ry2, px1, py1, px2, py2, fw, fh)

        # Always draw plate box
        cv2.rectangle(canvas, (fpx1, fpy1), (fpx2, fpy2), COLOR_PLATE, 2)

        if crop is None:
            draw_label(canvas, "Plate", fpx1, fpy1, COLOR_PLATE)
            continue

        raw = run_ocr(crop)
        if raw:
            cleaned = clean_plate(raw)
            if is_valid_plate(cleaned):
                _ocr_buffer[cleaned] = _ocr_buffer.get(cleaned, 0) + 1
                best_guess = max(_ocr_buffer, key=_ocr_buffer.get)
                draw_label(canvas, best_guess, fpx1, fpy1, COLOR_PLATE)
                if _ocr_buffer[cleaned] >= OCR_VOTE_WINDOW:
                    found_plates.append(cleaned)
                    _ocr_buffer.clear()
            else:
                draw_label(canvas, "Plate", fpx1, fpy1, COLOR_PLATE)
        else:
            draw_label(canvas, "Plate", fpx1, fpy1, COLOR_PLATE)

    return canvas, found_plates


def flush_ocr_buffer():
    if not _ocr_buffer:
        return
    best = max(_ocr_buffer, key=_ocr_buffer.get)
    if is_valid_plate(best):
        log_plate(best)
    _ocr_buffer.clear()


# ══════════════════════════════════════════════════════════════════════════════
# Save helpers
# ══════════════════════════════════════════════════════════════════════════════

def save_to_xlsx(ts, text):
    try:
        if os.path.exists(VIOLATIONS_XLSX):
            wb = openpyxl.load_workbook(VIOLATIONS_XLSX)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Violations"
            for col, h in enumerate(["#", "Timestamp", "Plate Number"], 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = PatternFill("solid", fgColor="1a1d24")
                cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions["A"].width = 6
            ws.column_dimensions["B"].width = 22
            ws.column_dimensions["C"].width = 20
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=ts)
        ws.cell(row=row, column=3, value=text)
        wb.save(VIOLATIONS_XLSX)
    except Exception as e:
        print(f"[WARN] XLSX: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# Display helper
# ══════════════════════════════════════════════════════════════════════════════

def frame_to_photoimage(bgr, cw, ch):
    fh, fw = bgr.shape[:2]
    scale  = min(cw / fw, ch / fh)
    nw, nh = max(1, int(fw*scale)), max(1, int(fh*scale))
    resized = cv2.resize(bgr, (nw, nh), interpolation=cv2.INTER_AREA)
    rgb     = cv2.cvtColor(resized, cv2.COLOR_BGR2RGB)
    return ImageTk.PhotoImage(image=Image.fromarray(rgb)), nw, nh


# ══════════════════════════════════════════════════════════════════════════════
# Detection loops
# ══════════════════════════════════════════════════════════════════════════════

def detection_loop():
    global running, cap, current_frame

    while running:

        # ── IMAGE ─────────────────────────────────────────────────────────────
        if source_type == "image" and source_path:
            img = cv2.imread(source_path)
            if img is None:
                print("[ERROR] Cannot read image.")
                break
            result, plates = process_frame_live(img)
            for p in plates:
                log_plate(p)
            with frame_lock:
                current_frame = result
            while running:
                time.sleep(0.1)
            break

        # ── VIDEO ─────────────────────────────────────────────────────────────
        elif source_type == "video" and source_path:
            cap = cv2.VideoCapture(source_path)
            if not cap.isOpened():
                print("[ERROR] Cannot open video.")
                break
            _fc = 0
            while running:
                ret, frame = cap.read()
                if not ret:
                    break
                _fc += 1
                if _fc % FRAME_SKIP != 0:
                    with frame_lock:
                        current_frame = frame
                    continue
                result, plates = process_frame_video(frame)
                for p in plates:
                    log_plate(p)
                with frame_lock:
                    current_frame = result
            flush_ocr_buffer()
            cap.release()
            break

        # ── WEBCAM ────────────────────────────────────────────────────────────
        else:
            cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
            if not cap.isOpened():
                cap = cv2.VideoCapture(0)
            if not cap.isOpened():
                print("[ERROR] Cannot open camera.")
                break
            cap.set(cv2.CAP_PROP_FRAME_WIDTH,  1280)
            cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
            while running:
                ret, frame = cap.read()
                if not ret or frame is None:
                    time.sleep(0.05)
                    continue
                result, plates = process_frame_live(frame)
                for p in plates:
                    log_plate(p)
                with frame_lock:
                    current_frame = result
                time.sleep(0.01)
            cap.release()
            break

    with frame_lock:
        current_frame = None
    root.after(0, lambda: update_ui_state(False))


# ══════════════════════════════════════════════════════════════════════════════
# Canvas refresh
# ══════════════════════════════════════════════════════════════════════════════

def refresh_canvas():
    global _last_photo
    if running:
        with frame_lock:
            frame = current_frame
        if frame is not None:
            cw = video_canvas.winfo_width()
            ch = video_canvas.winfo_height()
            if cw > 10 and ch > 10:
                try:
                    photo, nw, nh = frame_to_photoimage(frame, cw, ch)
                    _last_photo = photo
                    video_canvas.delete("all")
                    video_canvas.create_image(
                        (cw-nw)//2, (ch-nh)//2, anchor="nw", image=photo)
                except Exception:
                    pass
    root.after(30, refresh_canvas)


# ══════════════════════════════════════════════════════════════════════════════
# UI helpers
# ══════════════════════════════════════════════════════════════════════════════

def log_plate(text):
    global _seen_plates
    if not text or not is_valid_plate(text):
        return
    now = time.time()
    if text in _seen_plates and (now - _seen_plates[text]) < PLATE_COOLDOWN:
        return
    t_c = text.replace(" ", "")
    for seen, ts in list(_seen_plates.items()):
        if (now - ts) < PLATE_COOLDOWN:
            if edit_distance(t_c, seen.replace(" ", "")) <= 2:
                return
    _seen_plates[text] = now
    plate_log.append(text)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def _update():
        log_box.configure(state="normal")
        log_box.insert(tk.END, text + "\n")
        log_box.see(tk.END)
        log_box.configure(state="disabled")
    try:
        root.after(0, _update)
    except Exception:
        pass

    print(f"[PLATE] {text}")
    try:
        with open(VIOLATIONS_TXT, "a", encoding="utf-8") as f:
            f.write(f"{ts}  |  {text}\n")
    except Exception as e:
        print(f"[WARN] TXT: {e}")
    save_to_xlsx(ts, text)


def update_ui_state(is_running):
    if is_running:
        btn_start.configure(state="disabled", text="● Running")
        btn_stop.configure(state="normal")
        lbl_status.configure(text="● LIVE", fg="#1de9b6")
    else:
        btn_start.configure(state="normal", text="▶  Start")
        btn_stop.configure(state="disabled")
        lbl_status.configure(text="○ IDLE", fg="#6b7280")
        video_canvas.delete("all")
        cw = video_canvas.winfo_width()
        ch = video_canvas.winfo_height()
        video_canvas.create_text(cw//2, ch//2,
            text="Select a source and press Start",
            fill="#2e3140", font=("Courier New", 13))


def start_detection():
    global running, detection_thread, _ocr_buffer, _seen_plates
    if detection_thread and detection_thread.is_alive():
        return
    _ocr_buffer.clear()
    _seen_plates.clear()
    running = True
    update_ui_state(True)
    detection_thread = threading.Thread(target=detection_loop, daemon=True)
    detection_thread.start()


def stop_detection():
    global running, detection_thread
    running = False
    if detection_thread and detection_thread.is_alive():
        detection_thread.join(timeout=3.0)


def select_video():
    global source_type, source_path
    path = filedialog.askopenfilename(
        title="Select Video",
        filetypes=[("Video files", "*.mp4 *.avi *.mov *.mkv *.wmv"), ("All", "*.*")])
    if path:
        source_type = "video"
        source_path = path
        lbl_source.configure(text="📹  " + path.replace("\\", "/").split("/")[-1])


def select_image():
    global source_type, source_path
    path = filedialog.askopenfilename(
        title="Select Image",
        filetypes=[("Image files", "*.jpg *.jpeg *.png *.bmp *.webp"), ("All", "*.*")])
    if path:
        source_type = "image"
        source_path = path
        lbl_source.configure(text="🖼  " + path.replace("\\", "/").split("/")[-1])


def use_camera():
    global source_type, source_path
    source_type = "camera"
    source_path = None
    lbl_source.configure(text="📷  Webcam (live)")


def clear_log():
    plate_log.clear()
    log_box.configure(state="normal")
    log_box.delete("1.0", tk.END)
    log_box.configure(state="disabled")


# ══════════════════════════════════════════════════════════════════════════════
# UI Layout
# ══════════════════════════════════════════════════════════════════════════════

root = tk.Tk()
root.title("Helmet Violation Detector")
root.minsize(900, 560)
root.configure(bg="#0d0d0f")
root.resizable(True, True)

BG     = "#0d0d0f"
PANEL  = "#13151a"
PANEL2 = "#1a1d24"
ACCENT = "#00e5ff"
RED    = "#ff3c5f"
TEAL   = "#1de9b6"
BLUE   = "#448aff"
PURPLE = "#ab47bc"
MUTED  = "#6b7280"
TEXT   = "#dde1ea"

F_HEAD  = ("Courier New", 11, "bold")
F_LBL   = ("Courier New", 9)
F_BTN   = ("Courier New", 10, "bold")
F_LOG   = ("Courier New", 9)
F_TITLE = ("Courier New", 14, "bold")

topbar = tk.Frame(root, bg=BG, pady=10)
topbar.pack(fill="x", padx=16, pady=(10, 0))
tk.Label(topbar, text="⬡ HELMET VIOLATION DETECTOR",
         font=F_TITLE, fg=ACCENT, bg=BG).pack(side="left")
lbl_status = tk.Label(topbar, text="○ IDLE",
                      font=("Courier New", 10, "bold"), fg=MUTED, bg=BG)
lbl_status.pack(side="right", padx=6)
tk.Frame(root, bg=ACCENT, height=1).pack(fill="x", padx=16)

main = tk.Frame(root, bg=BG)
main.pack(fill="both", expand=True, padx=16, pady=10)
main.columnconfigure(1, weight=1)
main.rowconfigure(0, weight=1)

sidebar = tk.Frame(main, bg=PANEL, padx=14, pady=14, width=220)
sidebar.grid(row=0, column=0, sticky="ns", padx=(0, 10))
sidebar.pack_propagate(False)

def section(parent, title, color=ACCENT):
    tk.Label(parent, text=title, font=F_HEAD, fg=color, bg=PANEL).pack(anchor="w", pady=(8,2))
    tk.Frame(parent, bg=color, height=1).pack(fill="x", pady=(0,8))

def mkbtn(parent, text, cmd, color=ACCENT):
    b = tk.Button(parent, text=text, command=cmd,
                  font=F_BTN, fg="#0b0d10", bg=color,
                  activebackground=color, activeforeground="#0b0d10",
                  relief="flat", bd=0, cursor="hand2", pady=6)
    b.pack(fill="x", pady=2)
    return b

section(sidebar, "INPUT SOURCE")
mkbtn(sidebar, "📷  Webcam",      use_camera,   color=TEAL)
mkbtn(sidebar, "📹  Video File",  select_video, color=BLUE)
mkbtn(sidebar, "🖼   Image File",  select_image, color=PURPLE)

tk.Frame(sidebar, bg=PANEL, height=4).pack()
section(sidebar, "ACTIVE SOURCE")
lbl_source = tk.Label(sidebar, text="📷  Webcam (live)",
                      font=F_LBL, fg=TEXT, bg=PANEL,
                      wraplength=190, justify="left")
lbl_source.pack(anchor="w", pady=(0, 10))

section(sidebar, "CONTROLS")
btn_start = mkbtn(sidebar, "▶  Start", start_detection, color=ACCENT)
btn_stop  = mkbtn(sidebar, "■  Stop",  stop_detection,  color=RED)
btn_stop.configure(state="disabled")

tk.Frame(sidebar, bg=PANEL, height=8).pack()
tk.Label(sidebar, text="Saves to:\nviolations.txt\nviolations.xlsx",
         font=F_LBL, fg=MUTED, bg=PANEL, justify="left").pack(anchor="w")

centre = tk.Frame(main, bg=PANEL2)
centre.grid(row=0, column=1, sticky="nsew")
video_canvas = tk.Canvas(centre, bg="#0a0c10", highlightthickness=0)
video_canvas.pack(fill="both", expand=True)

right = tk.Frame(main, bg=PANEL, padx=12, pady=14, width=210)
right.grid(row=0, column=2, sticky="ns", padx=(10, 0))
right.pack_propagate(False)

section(right, "PLATE LOG", color=RED)
log_scroll = tk.Scrollbar(right, bg=PANEL, troughcolor=PANEL, relief="flat")
log_scroll.pack(side="right", fill="y")
log_box = tk.Text(right, font=F_LOG, bg="#080a0e", fg=TEAL,
                  insertbackground=ACCENT, selectbackground=RED,
                  relief="flat", bd=0, state="disabled",
                  yscrollcommand=log_scroll.set, wrap="word")
log_box.pack(side="left", fill="both", expand=True)
log_scroll.config(command=log_box.yview)
tk.Frame(right, bg=PANEL, height=6).pack()
mkbtn(right, "🗑  Clear", clear_log, color="#2a2d35")

tk.Frame(root, bg=ACCENT, height=1).pack(fill="x", padx=16)
btmbar = tk.Frame(root, bg=BG, pady=5)
btmbar.pack(fill="x", padx=16)
tk.Label(btmbar,
         text=f"two_wheeler.pt · helmet.pt · plate.pt · EasyOCR · {VIOLATIONS_TXT} · {VIOLATIONS_XLSX}",
         font=("Courier New", 8), fg=MUTED, bg=BG).pack(anchor="w")

def on_canvas_ready(event=None):
    video_canvas.delete("all")
    cw = video_canvas.winfo_width()
    ch = video_canvas.winfo_height()
    if cw > 10 and ch > 10:
        video_canvas.create_text(cw//2, ch//2,
            text="Select a source and press Start",
            fill="#2e3140", font=("Courier New", 13))

video_canvas.bind("<Configure>", on_canvas_ready)
root.after(100, refresh_canvas)
root.mainloop()